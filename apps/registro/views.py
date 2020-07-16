from django.shortcuts import render, redirect
from django.http import HttpResponseRedirect, HttpResponse
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView, TemplateView
from apps.registro.models import Registro
from apps.registro.forms import RegistroForm
from openpyxl import Workbook


def index_registro(request):
	return render(request, 'registro/index.html')


class RegistroList(ListView):
	model = Registro
	template_name = 'registro/index.html'


class RegistroCreate(CreateView):
	model = Registro
	form_class = RegistroForm
	template_name = 'registro/ingreso.html'
	success_url = reverse_lazy('registro:registro_listar')


class RegistroUpdate(UpdateView):
	model = Registro
	form_class = RegistroForm
	template_name = 'registro/ingreso.html'
	success_url = reverse_lazy('registro:registro_listar')


class RegistroDelete(DeleteView):
	model = Registro
	template_name = 'registro/eliminar.html'
	success_url = reverse_lazy('registro:registro_listar')

class RegistroDetail(DetailView):
	model = Registro
	template_name = "registro/detalle.html"

class ReporteRegistroExcel(TemplateView):
	
	def get(self,request,*args,**kwargs):
		registros = Registro.objects.all()
		wb = Workbook()
		ws = wb.active
		ws['B1'] = 'Reporte de registros biológicos de especies'
		ws.merge_cells('B1:G1')
		ws['B3'] = 'ESPECIE'
		ws['C3'] = 'FAMILIA'
		ws['D3'] = 'NOMBRE COMÚN'
		ws['E3'] = 'PROYECTO'
		ws['F3'] = 'BASE DEL REGISTRO'
		ws['G3'] = 'IDENTIFICADOR'
		ws['H3'] = 'AÑO IDENTIFICACIÓN'
		ws['I3'] = 'DEPARTAMENTO'
		ws['J3'] = 'MUNICIPIO'
		ws['K3'] = 'LOCALIDAD'
		ws['L3'] = 'LATITUD'
		ws['M3'] = 'LONGITUD'
		ws['N3'] = 'AUTOR'
		ws['O3'] = 'FECHA CAPTURA'
		ws['P3'] = 'ECORREGIÓN'

		cont = 4
		for reg in registros:
			ws.cell(row = cont, column = 2).value = reg.especie
			ws.cell(row = cont, column = 3).value = reg.familia
			ws.cell(row = cont, column = 4).value = reg.nombrecomun
			ws.cell(row = cont, column = 5).value = reg.proyecto.nombre
			ws.cell(row = cont, column = 6).value = reg.baseregistro
			ws.cell(row = cont, column = 7).value = reg.identificador
			ws.cell(row = cont, column = 8).value = reg.añoidentificacion
			ws.cell(row = cont, column = 9).value = reg.departamento
			ws.cell(row = cont, column = 10).value = reg.municipio
			ws.cell(row = cont, column = 11).value = reg.localidad
			ws.cell(row = cont, column = 12).value = reg.latitud
			ws.cell(row = cont, column = 13).value = reg.longitud
			ws.cell(row = cont, column = 14).value = reg.autor
			ws.cell(row = cont, column = 15).value = reg.fechacaptura
			ws.cell(row = cont, column = 16).value = reg.ecorregion.nombre

			cont+=1

			response = HttpResponse(content_type = "application/ms-excel")
			response['Content-Disposition'] = 'attachment; filename=ReporteEspecies.xlsx'
			wb.save(response)
			return response
