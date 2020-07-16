from django.shortcuts import render, redirect
from django.http import HttpResponseRedirect, HttpResponse
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, TemplateView
from apps.proyecto.models import Proyecto
from apps.proyecto.forms import ProyectoForm
from openpyxl import Workbook

def index_proyecto(request):
	return render(request, 'proyecto/index.html')


class ProyectoList(ListView):
	model = Proyecto
	template_name = 'Proyecto/index.html'


class ProyectoCreate(CreateView):
	model = Proyecto
	form_class = ProyectoForm
	template_name = 'proyecto/ingreso.html'
	success_url = reverse_lazy('proyecto:proyecto_listar')


class ProyectoUpdate(UpdateView):
	model = Proyecto
	form_class = ProyectoForm
	template_name = 'proyecto/ingreso.html'
	success_url = reverse_lazy('proyecto:proyecto_listar')


class ProyectoDelete(DeleteView):
	model = Proyecto
	template_name = 'proyecto/eliminar.html'
	success_url = reverse_lazy('proyecto:proyecto_listar')


class ReporteProyectoExcel(TemplateView):
	
	def get(self,request,*args,**kwargs):
		proyectos = Proyecto.objects.all()
		wb = Workbook()
		ws = wb.active
		ws['B1'] = 'Reporte de proyectos'
		ws.merge_cells('B1:G1')
		ws['B3'] = 'NOMBRE'
		ws['C3'] = 'INFORMACIÓN'
		

		cont = 4
		for pro in proyectos:
			ws.cell(row = cont, column = 2).value = pro.nombre
			ws.cell(row = cont, column = 3).value = pro.informacion
			cont = cont + 1

		response = HttpResponse(content_type = "application/ms-excel")
		response['Content-Disposition'] = 'attachment; filename=ReporteProyecto.xlsx'
		wb.save(response)
		return response