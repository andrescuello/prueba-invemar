from django.shortcuts import render, redirect
from django.http import HttpResponseRedirect, HttpResponse
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, TemplateView
from apps.ecorregion.models import Ecorregion
from apps.ecorregion.forms import EcorregionForm
from openpyxl import Workbook

def index_ecorregion(request):
	return render(request, 'ecorregion/index.html')


class EcorregionList(ListView):
	model = Ecorregion
	template_name = 'ecorregion/index.html'


class EcorregionCreate(CreateView):
	model = Ecorregion
	form_class = EcorregionForm
	template_name = 'Ecorregion/ingreso.html'
	success_url = reverse_lazy('ecorregion:ecorregion_listar')


class EcorregionUpdate(UpdateView):
	model = Ecorregion
	form_class = EcorregionForm
	template_name = 'ecorregion/ingreso.html'
	success_url = reverse_lazy('ecorregion:ecorregion_listar')


class EcorregionDelete(DeleteView):
	model = Ecorregion
	template_name = 'ecorregion/eliminar.html'
	success_url = reverse_lazy('ecorregion:ecorregion_listar')


class ReporteEcorregionExcel(TemplateView):
	
	def get(self,request,*args,**kwargs):
		ecorregiones = Ecorregion.objects.all()
		wb = Workbook()
		ws = wb.active
		ws['B1'] = 'Reporte de Ecorregión'
		ws.merge_cells('B1:G1')
		ws['B3'] = 'NOMBRE'
		

		cont = 4
		for eco in ecorregiones:
			ws.cell(row = cont, column = 2).value = eco.nombre
			cont = cont + 1

		response = HttpResponse(content_type = "application/ms-excel")
		response['Content-Disposition'] = 'attachment; filename=ReporteProyecto.xlsx'
		wb.save(response)
		return response